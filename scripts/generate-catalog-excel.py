import json
import io
import base64
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

BASE = Path(__file__).resolve().parent.parent
MANIFEST = BASE / "public" / "manifest.json"
IMAGES_BASE = BASE / "public" / "images" / "Categorias"
OUTPUT_XLSX = BASE / "catalogo_productos.xlsx"

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
ROW_HEIGHT_PX = 280  # altura de cada fila con imagen
ROW_HEIGHT_POINTS = ROW_HEIGHT_PX * 0.75  # openpyxl usa puntos (1px = 0.75pt)
IMG_COL_WIDTH = 48  # ancho columna imagen (unidades de caracter)
JPEG_QUALITY = 88


def parse_metadata(path: Path) -> dict:
    data = {}
    if not path.exists():
        return data
    text = path.read_text(encoding="utf-8")
    for line in text.splitlines():
        if ":" not in line:
            continue
        key, _, value = line.partition(":")
        data[key.strip().lower()] = value.strip()
    return data


def find_first_image(product_dir: Path, meta: dict) -> Path | None:
    # Prioridad: primero listado en metadata; si no, primer archivo del dir
    images_field = meta.get("images", "")
    if images_field:
        first = images_field.split(",")[0].strip()
        if first:
            candidate = product_dir / first
            if candidate.exists():
                return candidate
    if product_dir.exists():
        for f in sorted(product_dir.iterdir()):
            if f.suffix.lower() in IMAGE_EXTS:
                return f
    return None


def image_to_xl(img_path: Path, max_h_px: int = ROW_HEIGHT_PX - 6) -> XLImage | None:
    """Convierte cualquier formato soportado a JPEG en memoria y lo redimensiona."""
    try:
        with PILImage.open(img_path) as im:
            im = im.convert("RGB")
            w, h = im.size
            if h > max_h_px:
                ratio = max_h_px / h
                new_w = max(1, int(w * ratio))
                im = im.resize((new_w, max_h_px), PILImage.LANCZOS)
            buf = io.BytesIO()
            im.save(buf, format="JPEG", quality=JPEG_QUALITY, optimize=True)
            buf.seek(0)
            return XLImage(buf)
    except Exception as e:
        print(f"  ! Error con imagen {img_path}: {e}")
        return None


def main():
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Catalogo"

    headers = [
        "Codigo",
        "Imagen",
        "Categoria",
        "Producto",
        "Descripcion",
        "Precio Mayorista",
        "Precio Minorista",
        "Proveedor",
        "Activo",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total = 0
    images_embedded = 0
    missing_images = []

    def to_num(v):
        if v == "" or v is None:
            return ""
        try:
            return int(v)
        except ValueError:
            try:
                return float(v)
            except ValueError:
                return v

    # Set de (categoria, producto) que figuran en manifest.json para el flag "activo"
    manifest_set = {
        (cat, prod) for cat, prods in manifest.items() for prod in prods
    }

    # Recolectamos TODOS los productos presentes en el filesystem.
    # Activo = "si" si aparece en manifest.json, "no" si no.
    all_products = []  # list of (category, product_name)
    for cat_dir in sorted(IMAGES_BASE.iterdir()):
        if not cat_dir.is_dir():
            continue
        for prod_dir in sorted(cat_dir.iterdir()):
            if not prod_dir.is_dir():
                continue
            all_products.append((cat_dir.name, prod_dir.name))

    row_idx = 2
    for category, product_name in all_products:
            product_dir = IMAGES_BASE / category / product_name
            meta = parse_metadata(product_dir / "metadata.txt")

            code = meta.get("code", "")
            description = meta.get("description", "")
            price_mayorista = meta.get("price_mayorista", "")
            price_minorista = meta.get("price_minorista", "")
            proveedor = meta.get("proveedor", "")
            activo = "si" if (category, product_name) in manifest_set else "no"

            ws.cell(row=row_idx, column=1, value=code)
            # columna 2 queda para la imagen
            ws.cell(row=row_idx, column=3, value=category)
            ws.cell(row=row_idx, column=4, value=product_name)
            ws.cell(row=row_idx, column=5, value=description)
            ws.cell(row=row_idx, column=6, value=to_num(price_mayorista))
            ws.cell(row=row_idx, column=7, value=to_num(price_minorista))
            ws.cell(row=row_idx, column=8, value=proveedor)
            ws.cell(row=row_idx, column=9, value=activo)

            img_path = find_first_image(product_dir, meta)
            if img_path is not None:
                xl_img = image_to_xl(img_path)
                if xl_img is not None:
                    anchor = f"B{row_idx}"
                    ws.add_image(xl_img, anchor)
                    images_embedded += 1
                else:
                    missing_images.append(f"{category}/{product_name}")
            else:
                missing_images.append(f"{category}/{product_name}")

            ws.row_dimensions[row_idx].height = ROW_HEIGHT_POINTS
            total += 1
            row_idx += 1

    # Header row height
    ws.row_dimensions[1].height = 22

    # Column widths (A..I)
    widths = [14, IMG_COL_WIDTH, 18, 32, 80, 18, 18, 16, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for col_letter in ("F", "G"):
        for cell in ws[col_letter][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    wb.save(OUTPUT_XLSX)

    print(f"Productos escritos: {total}")
    print(f"Imagenes embebidas: {images_embedded}")
    if missing_images:
        print(f"Sin imagen ({len(missing_images)}):")
        for m in missing_images:
            print(f"  - {m}")
    print(f"Archivo: {OUTPUT_XLSX}")

    with open(OUTPUT_XLSX, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("ascii")
    b64_path = OUTPUT_XLSX.with_suffix(".b64")
    b64_path.write_text(b64)
    print(f"Base64: {b64_path} ({len(b64)} chars)")


if __name__ == "__main__":
    main()
